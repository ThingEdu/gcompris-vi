#!/usr/bin/env python3
"""Sinh bảng đối chiếu chuẩn cho bản GCompris tiếng Việt của ThingEdu.

Theo đúng mẫu ThingEdu_DoiChieu_Chuan_GDPT2018_NLS2025_NGSS_K1K12: thang mức
liên kết 2/3/5, mã màu, phông Arial, cột Minh chứng.

Usage: make_xlsx.py <acts.json> <out.xlsx>
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BLUE = "FF1340FF"
ROW_BG = "FFE8EDFF"
M5, M3, M2 = "FFC9E7CB", "FFFFF2CC", "FFE3E3E3"
GREY = "FFF3F3F3"
THIN = Side(style="thin", color="FFD5DCF5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def title(ws, text, sub, ncol):
    ws["A1"] = text
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color=BLUE)
    ws["A2"] = sub
    ws["A2"].font = Font(name="Arial", size=9, italic=True, color="FF666666")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)


def legend(ws, row, ncol):
    ws.cell(row, 1, "Căn cứ xếp mức").font = Font(name="Arial", size=9, bold=True)
    ws.cell(row, ncol, "Mã màu").font = Font(name="Arial", size=9, bold=True)
    txt = [
        ("Chưa có bằng chứng chuẩn này được đáp ứng trong bản GCompris tiếng Việt.", 2, M2),
        ("Chuẩn CÓ THỂ được đáp ứng — tùy cách Mentor tổ chức buổi học quanh hoạt động.", 3, M3),
        ("Chuẩn được đáp ứng RÕ RÀNG bởi chính hoạt động trong phần mềm.", 5, M5),
    ]
    for i, (t, lv, fill) in enumerate(txt, 1):
        ws.cell(row + i, 1, t).font = Font(name="Arial", size=9)
        ws.merge_cells(start_row=row + i, start_column=1, end_row=row + i, end_column=ncol - 1)
        c = ws.cell(row + i, ncol, lv)
        c.fill = PatternFill("solid", start_color=fill)
        c.alignment = Alignment(horizontal="center")
        c.font = Font(name="Arial", size=9)
    return row + 5


def header(ws, row, cols):
    for i, h in enumerate(cols, 1):
        c = ws.cell(row, i, h)
        c.fill = PatternFill("solid", start_color=BLUE)
        c.font = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.border = BORDER
    ws.row_dimensions[row].height = 32
    ws.freeze_panes = ws.cell(row + 1, 1)


def rows(ws, start, data, level_col=None, bold_col=None, stripe=True):
    for r, vals in enumerate(data, start):
        for i, v in enumerate(vals, 1):
            c = ws.cell(r, i, v)
            c.font = Font(name="Arial", size=9, bold=(i == bold_col))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER
            if level_col and i == level_col:
                c.fill = PatternFill("solid", start_color={5: M5, 3: M3, 2: M2}.get(v, ROW_BG))
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif stripe:
                c.fill = PatternFill("solid", start_color=ROW_BG if r % 2 else GREY)
    return start + len(data)


def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w


# ---------------------------------------------------------------- phân loại
MICROWORLD = {
    "analog_electricity", "digital_electricity", "gravity", "land_safe", "submarine",
    "watercycle", "renewable_energy", "solar_system", "canal_lock", "balancebox",
    "binary_bulb", "color_mix", "color_mix_light", "programmingMaze",
    "path_encoding", "path_decoding", "path_encoding_relative", "path_decoding_relative",
    "scalesboard", "scalesboard_weight", "scalesboard_weight_avoirdupois", "share",
    "learn_decimals", "learn_decimals_additions", "learn_decimals_subtractions",
    "learn_quantities", "fractions_create", "fractions_find",
    "graduated_line_read", "graduated_line_use", "frieze", "graph_coloring",
    "mosaic", "clockgame",
}
TOOL = {
    "sketch", "simplepaint", "piano_composition", "baby_wordprocessor",
    "drawing_wheels", "compass", "clickanddraw", "redraw", "redraw_symmetrical",
}
OBJECT = {
    "chess", "chess_2players", "chess_partyend", "checkers", "checkers_2players",
    "align4", "align4_2players", "bargame", "bargame_2players",
    "nine_men_morris", "nine_men_morris_2players", "oware", "oware_2players",
    "tic_tac_toe", "tic_tac_toe_2players", "hanoi", "hanoi_real", "tangram",
    "baby_tangram", "fifteen", "traffic", "sudoku", "superbrain", "lightsoff",
    "penalty", "maze", "mazeinvisible", "mazerelative", "hexagon", "guess24",
    "guesscount", "guessnumber", "calcudoku", "crane", "railroad", "algorithm",
    "babyshapes", "babymatch", "photo_hunter", "details", "paintings", "chronos",
}
ROLE_NOTE = {
    "Vi thế giới": "Trẻ thao tác trên một thế giới nhỏ có luật riêng, thử giả thuyết rồi nhận phản hồi ngay — đúng lõi Papert.",
    "Công cụ sáng tạo": "Trẻ làm ra sản phẩm của mình chứ không trả lời câu hỏi của máy.",
    "Đồ vật để suy nghĩ": "Bài toán cụ thể cầm nắm được, dùng để nghĩ về chiến thuật và cấu trúc.",
    "Luyện tập có phản hồi": "Máy hỏi – trẻ đáp. Papert xếp đây là mức dùng máy tính yếu nhất; bắt buộc phải nối sang việc làm tay.",
}

# Thứ tự có ý nghĩa: nhóm hẹp xét trước nhóm rộng. GCompris gán 'sciences' cho cả
# hoạt động địa lí và lịch sử, nên hai nhóm đó phải đứng trước; 'braille' đứng
# trước 'reading' vì hoạt động chữ nổi mang cả hai nhãn.
SUBJ = [
    (("braille",), "Hoạt động trải nghiệm (giáo dục hòa nhập)"),
    (("geography", "history"), "Lịch sử và Địa lí"),
    (("letters", "words", "vocabulary", "reading"), "Tiếng Việt"),
    (("arithmetic", "numeration", "measures", "math", "money", "addition",
      "subtraction", "multiplication", "division"), "Toán"),
    (("sciences", "experiment"), "Tự nhiên và Xã hội (TH) · Khoa học · KHTN"),
    (("computer", "keyboard", "mouse", "mobile"), "Tin học"),
    (("arts",), "Mĩ thuật"),
    (("music",), "Âm nhạc"),
]
NLS = [
    (("computer", "keyboard", "mouse", "mobile"), "5.1 Giải quyết vấn đề kỹ thuật"),
    (("discovery", "sciences", "experiment"), "1.1 Duyệt, tìm kiếm và lọc dữ liệu"),
    (("arts", "music"), "3.1 Phát triển nội dung số"),
    (("logic", "strategy", "puzzle", "memory", "fun", "maze"), "5.3 Sử dụng sáng tạo công nghệ số"),
]
AGE = {1: "2–4 tuổi (mầm non)", 2: "4–6 tuổi (mầm non – K1)", 3: "5–7 tuổi (K1–K2)",
       4: "7–8 tuổi (K2–K3)", 5: "8–9 tuổi (K3–K4)", 6: "9–10 tuổi (K4–K5)"}
MAKER = {
    "menu": "Không phải hoạt động học — đây là màn hình chính để trẻ tự chọn.",
    "canal_lock": "Âu tàu bằng hai hộp nhựa nối ống: đổ nước, mở van, xem thuyền giấy lên xuống theo mực nước.",
    "binary_bulb": "Tám cái đèn pin hoặc tám tấm bìa lật, mỗi cái một giá trị 1-2-4-8… Trẻ bật tắt để đếm.",
    "land_safe": "Thả quả bóng từ các độ cao khác nhau, bấm giờ và ghi lại — cảm nhận gia tốc trước.",
    "submarine": "Chai nhựa có lỗ và ống hút làm tàu ngầm cartesian: bóp chai thì chìm, thả ra thì nổi.",
    "frieze": "Dải hoa văn bằng nắp chai, hạt và giấy màu dán trên băng giấy dài quanh lớp.",
    "graduated_line_read": "Thước mét dán lên tường, trẻ chỉ và đọc vạch.",
    "graduated_line_use": "Thước mét dán lên tường, trẻ dán kẹp giấy vào đúng vạch được đọc lên.",
    "learn_decimals": "Mười que tính bó thành một bó — một bó là một đơn vị, một que là một phần mười.",
    "learn_quantities": "Rổ cam, quýt thật để trẻ đếm và bốc đúng số lượng.",
    "path_decoding": "Bạn A đọc dãy lệnh, bạn B bịt mắt đi theo trong lớp.",
    "path_decoding_relative": "Bạn A đọc lệnh quay trái, quay phải; bạn B bịt mắt làm theo — hướng tính theo thân mình.",
    "path_encoding": "Trẻ viết dãy lệnh ra thẻ giấy để bạn khác đi theo mà tới đúng chỗ.",
    "path_encoding_relative": "Trẻ viết lệnh quay trái, quay phải ra thẻ giấy cho bạn bịt mắt đi theo.",
    "analog_electricity": "Lắp mạch thật: pin, bóng, dây, công tắc — so kết quả với mô phỏng.",
    "digital_electricity": "Cổng logic bằng công tắc và bóng đèn trên bảng gỗ.",
    "scalesboard": "Cân đĩa thật với quả cân tự làm từ đồng xu, hạt.",
    "watercycle": "Mô hình vòng tuần hoàn nước bằng chai nhựa, đèn và đá lạnh.",
    "renewable_energy": "Tua-bin gió giấy và pin mặt trời mini thắp một bóng LED.",
    "solar_system": "Mô hình hệ Mặt Trời tỉ lệ bằng bóng nhựa trên sân trường.",
    "programmingMaze": "Thẻ lệnh giấy điều khiển bạn bịt mắt đi trong lớp, trước khi lập trình trên máy.",
    "smallnumbers": "Xúc xắc thật, ném và đếm.",
    "smallnumbers2": "Bộ đô-mi-nô thật.",
    "clockgame": "Đồng hồ giấy có kim xoay được, tự làm.",
    "tangram": "Bộ Tangram cắt từ bìa cứng.",
    "hanoi_real": "Tháp Hà Nội bằng đĩa giấy hoặc nắp chai.",
    "share": "Chia kẹo thật giữa các bạn trong lớp.",
    "fractions_create": "Cắt bánh giấy thành phần bằng nhau.",
    "money": "Tiền giấy in cho lớp học — xem lưu ý trong sheet CanThietKeLai.",
    "balancebox": "Mê cung nghiêng bằng bìa và bi ve.",
    "compass": "Compa thật trên giấy.",
    "sketch": "Vẽ tay trước, số hóa sau.",
    "piano_composition": "Đàn phím thật hoặc bộ ống nước cắt theo cao độ.",
    "gravity": "Thả vật rơi và bấm giờ ngoài sân.",
    "melody": "Bộ chai nước gõ theo cao độ.",
    "crane": "Cần cẩu giấy điều khiển bằng ròng rọc.",
    "railroad": "Xếp lại đoàn tàu bằng khối gỗ.",
}
MAKER_BY_SECTION = {
    "math": "Đếm và đo bằng vật thật trước khi lên màn hình.",
    "arithmetic": "Que tính, hạt, nắp chai — làm phép tính bằng tay trước.",
    "letters": "Thẻ chữ cắt rời, ghép trên bàn.",
    "words": "Thẻ chữ và tranh, ghép trên bàn.",
    "reading": "Đọc to thành tiếng cùng bạn, rồi mới đối chiếu trên máy.",
    "memory": "Bộ thẻ lật tự làm bằng bìa.",
    "sciences": "Làm lại thí nghiệm bằng vật liệu thật trong Làng.",
    "experiment": "Làm lại thí nghiệm bằng vật liệu thật trong Làng.",
    "geography": "Bản đồ ghép bằng bìa, dán lên tường lớp.",
    "history": "Dòng thời gian bằng dây và kẹp ảnh treo ngang lớp.",
    "music": "Nhạc cụ tự chế: chai nước, ống nhựa, hộp dây thun.",
    "arts": "Vẽ, cắt, dán trên giấy thật.",
    "puzzle": "Bộ xếp hình cắt từ bìa.",
    "strategy": "Bàn cờ vẽ tay, quân bằng nắp chai.",
    "logic": "Vật thật xếp theo quy luật trên bàn.",
    "computer": "Bàn phím cũ tháo rời cho trẻ sờ và gọi tên phím.",
    "mouse": "Bàn phím cũ tháo rời cho trẻ sờ và gọi tên phím.",
    "keyboard": "Bàn phím cũ tháo rời cho trẻ sờ và gọi tên phím.",
    "measures": "Cân, thước, cốc đong thật.",
    "money": "Tiền giấy in cho lớp học.",
    "color": "Pha màu nước thật trong đĩa.",
    "colors": "Pha màu nước thật trong đĩa.",
    "braille": "Bảng chữ nổi dập bằng đinh ghim trên bìa.",
}


def role_of(a):
    if a["name"] in MICROWORLD:
        return "Vi thế giới"
    if a["name"] in TOOL:
        return "Công cụ sáng tạo"
    if a["name"] in OBJECT:
        return "Đồ vật để suy nghĩ"
    return "Luyện tập có phản hồi"


# Vài hoạt động mang nhãn chủ đề gây hiểu nhầm: mục đích thật là làm quen thiết
# bị chứ không phải học chữ. Ghi đè tay cho đúng môn.
SUBJ_OVERRIDE = {
    "baby_keyboard": "Tin học",
    "baby_wordprocessor": "Tin học",
    "keyboard_training": "Tin học",
    "ballcatch": "Tin học",
    "menu": "— (màn hình chính, không phải hoạt động học)",
}


def subject_of(a):
    if a["name"] in SUBJ_OVERRIDE:
        return SUBJ_OVERRIDE[a["name"]]
    secs = a["section"].split()
    for keys, name in SUBJ:
        if any(s in keys for s in secs):
            return name
    return "Hoạt động trải nghiệm"


def nls_of(a):
    secs = a["section"].split()
    for keys, name in NLS:
        if any(s in keys for s in secs):
            return name
    return "3.4 Lập trình" if "computer" in secs else "5.3 Sử dụng sáng tạo công nghệ số"


# Xét nhóm hẹp trước nhóm rộng, cùng lý do như SUBJ: GCompris gán 'sciences' cho
# cả hoạt động địa lí và lịch sử.
MAKER_ORDER = ["braille", "geography", "history", "money", "color", "colors", "music",
               "arts", "letters", "words", "reading", "measures", "arithmetic",
               "computer", "keyboard", "mouse", "memory", "puzzle", "strategy",
               "logic", "sciences", "experiment", "math"]


def maker_of(a):
    if a["name"] in MAKER:
        return MAKER[a["name"]]
    secs = set(a["section"].split())
    for key in MAKER_ORDER:
        if key in secs and key in MAKER_BY_SECTION:
            return MAKER_BY_SECTION[key]
    return "Nối sang một việc làm bằng tay cùng chủ đề trong Làng."


def build(acts, out_path):
    acts = [a for a in acts if a["name"] != "template"]
    wb = Workbook()

    # ------------------------------------------------------------ HuongDan
    ws = wb.active
    ws.title = "HuongDan"
    title(ws, "GCOMPRIS TIẾNG VIỆT — BẢNG ĐỐI CHIẾU CHUẨN",
          "ThingEdu · Làng Maker — đối chiếu GDPT 2018, Khung năng lực số TT 02/2025, "
          "Công văn 3456/BGDĐT-GDPT, dưới lăng kính Papert và Maker", 3)
    lines = [
        ("PHẦN MỀM ĐƯỢC ĐỐI CHIẾU", ""),
        ("GCompris tiếng Việt (ThingEdu)",
         "Bản việt hóa GCompris 26.1 — 202 hoạt động giáo dục cho trẻ 2–10 tuổi, cộng đồng KDE, giấy phép AGPL v3. "
         "Bản dịch 4.277 chuỗi + kho giọng đọc tiếng Việt sinh bằng VieNeu-TTS. "
         "Mã nguồn: github.com/ThingEdu/gcompris-vi. Chạy trên NEO One (Linux ARM64)."),
        ("BỐN CĂN CỨ ĐỐI CHIẾU", ""),
        ("1) GDPT 2018 — TT 32/2018/TT-BGDĐT",
         "Hệ thống môn học cấp Tiểu học; 5 phẩm chất, 3 năng lực chung, 7 năng lực đặc thù. "
         "Xem sheet GDPT2018-MonHoc và GDPT2018-NangLuc."),
        ("2) Khung năng lực số — TT 02/2025/TT-BGDĐT",
         "6 miền, 24 năng lực thành phần, 8 bậc. Hiệu lực 11/03/2025. Xem sheet NLS-TT02-2025."),
        ("3) Công văn 3456/BGDĐT-GDPT ngày 27/6/2025",
         "Hướng dẫn triển khai khung năng lực số cho học sinh phổ thông từ năm học 2025–2026: "
         "không lập môn riêng mà lồng ghép vào môn học hiện có, môn Tin học giữ vai trò nòng cốt, "
         "tích hợp qua hoạt động giáo dục – trải nghiệm – câu lạc bộ, không gây quá tải. "
         "Xem sheet CV3456-TrienKhai."),
        ("4) Tư duy Papert (constructionism)",
         "Seymour Papert: trẻ học tốt nhất khi tự tay làm ra một thứ có ý nghĩa với mình. "
         "Bốn khái niệm dùng trong bảng này: vi thế giới (microworld), đồ vật để suy nghĩ "
         "(objects-to-think-with), gỡ lỗi thay vì chấm điểm sai, và sàn thấp – trần cao – tường rộng. "
         "Xem sheet Papert-Maker."),
        ("CÁCH DÙNG", ""),
        ("Đối tượng sử dụng",
         "Đối thoại với trường học và phòng giáo dục (chứng minh GCompris tiếng Việt bổ trợ GDPT 2018 và "
         "phục vụ nhiệm vụ NLS theo CV 3456), hồ sơ đối tác, và định hướng cho Mentor khi soạn buổi học."),
        ("Cột 'Minh chứng'",
         "Nêu hoạt động cụ thể trong GCompris tiếng Việt chứng minh mức liên kết. "
         "Tên hoạt động ghi theo đúng bản dịch tiếng Việt đang chạy."),
        ("Điều KHÔNG hứa",
         "GCompris là phần mềm luyện tập và khám phá trên màn hình. Nó KHÔNG thay được giờ làm tay, "
         "không có nội dung an toàn số, không có nội dung AI. Ba miền đó phải đến từ chương trình The Lab "
         "và Thing Notebook. Sheet NLS-TT02-2025 ghi rõ mức 2 ở những chuẩn này."),
        ("MÃ MÀU MỨC LIÊN KẾT (theo đúng thang FLL, giống bộ đối chiếu The Lab)", ""),
        ("2", "Chưa có bằng chứng chuẩn được đáp ứng trong phần mềm"),
        ("3", "Có thể đạt, tùy cách Mentor tổ chức buổi học quanh hoạt động"),
        ("5", "Đạt rõ ràng qua chính hoạt động trong phần mềm"),
        ("CHÍN SHEET TRONG FILE", ""),
        ("TongHop", "Bảng tổng hợp số chuẩn theo từng mức, tính tự động từ bốn sheet đối chiếu."),
        ("GDPT2018-MonHoc · GDPT2018-NangLuc · NLS-TT02-2025 · CV3456-TrienKhai",
         "Bốn bảng đối chiếu tổng hợp, mỗi dòng một chuẩn."),
        ("Papert-Maker",
         "Phân loại 202 hoạt động theo vai trò trong tư duy Papert, kèm việc làm tay đi kèm."),
        ("HoatDong-202",
         "Ma trận đầy đủ: từng hoạt động gắn môn GDPT 2018, năng lực số, vai trò Papert, "
         "độ tuổi và việc nối tay."),
        ("CanThietKeLai",
         "Bốn chỗ nội dung không dịch được mà phải làm lại cho Việt Nam — đọc trước khi đưa vào lớp."),
    ]
    r = 4
    for a, b in lines:
        ca = ws.cell(r, 1, a)
        ca.font = Font(name="Arial", size=10, bold=not b, color=BLUE if not b else "FF000000")
        ca.alignment = Alignment(wrap_text=True, vertical="top")
        if b:
            cb = ws.cell(r, 2, b)
            cb.font = Font(name="Arial", size=9)
            cb.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        r += 1
    widths(ws, {"A": 42, "B": 78, "C": 20})

    # -------------------------------------------------------- GDPT môn học
    ws = wb.create_sheet("GDPT2018-MonHoc")
    title(ws, "ĐỐI CHIẾU HỆ THỐNG MÔN HỌC & HOẠT ĐỘNG GIÁO DỤC CẤP TIỂU HỌC — GDPT 2018",
          "GCompris tiếng Việt (ThingEdu) — 202 hoạt động, trẻ 2–10 tuổi", 5)
    r = legend(ws, 4, 5)
    header(ws, r, ["Môn học / HĐGD (GDPT 2018)", "Mức liên kết", "Số hoạt động GCompris",
                   "Minh chứng — hoạt động tiêu biểu trong bản tiếng Việt", "Ghi chú phạm vi"])
    cnt = {}
    for a in acts:
        cnt.setdefault(subject_of(a), []).append(a)

    def ex(subject, n=4):
        return " • ".join(x["vi_title"] for x in cnt.get(subject, [])[:n] if x["vi_title"])

    n = {k: len(v) for k, v in cnt.items()}
    n2p = sum(1 for a in acts if "2players" in a["name"] or a["name"].endswith("_tux")
              or "tux" in a["name"].lower())
    mon = [
        ("Toán", 5, "Phép cộng · Phép trừ · Phép nhân · Phép chia · Học số thập phân · Tạo phân số · "
                    "Cân theo hệ đo lường quốc tế · Đọc trục số có vạch chia · Phép cộng đặt tính dọc",
         f"Mạnh nhất: {n.get('Toán', 0)} hoạt động thuộc mạch số học – số – đo lường, phủ gần trọn mạch Toán Tiểu học."),
        ("Tiếng Việt", 3, "Chữ cái rơi · Từ rơi · Trò đoán chữ · Chữ cái còn thiếu · Sắp câu cho đúng · "
                          "Thứ tự bảng chữ cái",
         "Bảng chữ cái đã việt hóa đủ ă â đ ê ô ơ ư. NHƯNG bộ từ theo cấp độ âm tiết cho Chữ cái rơi và "
         "Từ rơi chưa soạn — xem sheet CanThietKeLai. Chưa dùng thay giờ Tiếng Việt được."),
        ("Tự nhiên và Xã hội (TH) · Khoa học · KHTN", 5,
         "Vòng tuần hoàn nước · Năng lượng tái tạo · Hệ Mặt Trời · Mạch điện tương tự · Mạch điện số · "
         "Trọng lực · Khám phá vật nuôi trong trang trại · Pha màu sơn",
         f"{n.get('Tự nhiên và Xã hội (TH) · Khoa học · KHTN', 0)} hoạt động khoa học – thí nghiệm. Đây là chỗ GCompris gần OpenSciEd nhất."),
        ("Tin học", 5, "Bàn phím đầu tiên của bé · Luyện gõ bàn phím · Luyện bấm chuột · Lập trình đường đi · "
                       "Mã hoá đường đi · Bóng đèn nhị phân",
         "Đúng vai trò nòng cốt mà CV 3456 giao cho môn Tin học ở tiểu học."),
        ("Lịch sử và Địa lí", 3, "Tìm quốc gia trên bản đồ · Tìm vùng trên bản đồ · Dòng thời gian · "
                                 "Sắp sự kiện theo thời gian",
         "CHƯA CÓ BẢN ĐỒ VIỆT NAM. Bản đồ hiện có là châu Âu, Mỹ, Ấn Độ, Trung Quốc, Úc. "
         "Phải bổ sung bộ 34 tỉnh thành thì mới dùng cho phần Địa lí Việt Nam."),
        ("Mĩ thuật", 5, "Vẽ tự do · Vẽ bằng bánh răng · Vẽ trên lưới ô · Vẽ hình đối xứng · Ghép tranh",
         "Có cả công cụ vẽ tự do lẫn bài tập tái tạo hình. Tranh dùng làm mẫu là tranh kinh điển thế giới."),
        ("Âm nhạc", 5, "Chơi dương cầm · Sáng tác dương cầm · Gõ nhịp · Gọi tên nốt nhạc · Nhạc cụ · Giai điệu",
         "Tên nốt đã việt hóa theo Đô Rê Mi. Chưa có dân ca Việt Nam trong bộ giai điệu mẫu."),
        ("Hoạt động trải nghiệm", 5, "Cờ vua · Cờ đam · Tháp Hà Nội · Trò xếp hình Tangram · Sudoku · "
                                     "Mê cung · Đá phạt đền",
         "Nhóm trò chơi tư duy và vận động tinh — hợp với giờ câu lạc bộ mà CV 3456 khuyến khích."),
        ("Đạo đức", 2, "—", "GCompris không có nội dung đạo đức, giá trị sống. Phần này thuộc về The Lab."),
        ("Giáo dục thể chất", 2, "—", "Không có nội dung vận động thô. Bù bằng NeoArcade và NeoAiSport."),
        ("Hoạt động trải nghiệm (giáo dục hòa nhập)", 5, "Khám phá hệ chữ nổi Braille · Vui cùng chữ nổi",
         "Hai hoạt động dạy hệ chữ nổi Braille — hiếm có trong phần mềm giáo dục tiếng Việt. "
         "Dùng được cho lớp hòa nhập và cho hoạt động hiểu bạn khiếm thị."),
        ("Ngoại ngữ (Tiếng Anh)", 3, "Mở rộng vốn từ",
         "Hoạt động Mở rộng vốn từ cho học song ngữ, nhưng 564 tệp giọng từ vựng chưa sinh — chưa dùng được."),
    ]
    rows_data = [(m, lv, str(len(cnt.get(m, []))) if m in cnt else "0", exs, note)
                 for m, lv, exs, note in mon]
    r = rows(ws, r + 1, rows_data, level_col=2, bold_col=1)
    ws.cell(r + 1, 1, "Ghi chú: cột 'Số hoạt động' đếm theo nhóm chủ đề mà GCompris tự gán cho hoạt động; "
                      "một hoạt động chỉ được đếm một lần.").font = Font(name="Arial", size=8, italic=True)
    widths(ws, {"A": 34, "B": 11, "C": 13, "D": 58, "E": 52})
    return wb, acts, cnt


def build_rest(wb, acts, cnt):
    n = {k: len(v) for k, v in cnt.items()}
    # ------------------------------------------------ GDPT phẩm chất năng lực
    ws = wb.create_sheet("GDPT2018-NangLuc")
    title(ws, "ĐỐI CHIẾU 5 PHẨM CHẤT — 3 NĂNG LỰC CHUNG — 7 NĂNG LỰC ĐẶC THÙ — GDPT 2018",
          "GCompris tiếng Việt (ThingEdu)", 4)
    r = legend(ws, 4, 4)
    header(ws, r, ["Nhóm", "Phẩm chất / Năng lực", "Mức liên kết",
                   "Minh chứng trong bản GCompris tiếng Việt"])
    d = [
        ("5 phẩm chất", "Yêu nước", 2, "Không có nội dung. Bản đồ Việt Nam và dân ca Việt còn thiếu — nếu bổ sung thì mức này mới đổi."),
        ("5 phẩm chất", "Nhân ái", 3, "Hoạt động Chia kẹo đặt bài toán chia đều cho các bạn; Gia đình dạy xưng hô người thân. Mentor phải gọi tên giá trị thì mới thành phẩm chất."),
        ("5 phẩm chất", "Chăm chỉ", 5, "Mọi hoạt động đều nhiều cấp, sai thì làm lại chứ không bị chấm điểm. Thanh sao vàng–đỏ cho trẻ thấy đường tiến."),
        ("5 phẩm chất", "Trung thực", 3, "Phần mềm tự chấm nên không có chỗ gian; nhưng chuẩn mực trung thực phải do Mentor đặt ra trong giờ."),
        ("5 phẩm chất", "Trách nhiệm", 3, "Chế độ nhiều người chơi và các trò đấu với bạn tạo tình huống giữ luật. Cần Mentor tổ chức."),
        ("3 năng lực chung", "Tự chủ và tự học", 5, "Trẻ tự chọn hoạt động trên menu, tự chọn cấp độ, tự thử lại. Không có giới hạn thời gian hay điểm số ép."),
        ("3 năng lực chung", "Giao tiếp và hợp tác", 3, "Có hoạt động chơi cặp và đấu với Tux. Hợp tác thật thì cần Mentor ghép cặp và tổ chức."),
        ("3 năng lực chung", "Giải quyết vấn đề và sáng tạo", 5, "Lập trình đường đi, Mạch điện tương tự, Mạch điện số, Tháp Hà Nội, Cờ chốt, Vẽ tự do — trẻ thử, sai, sửa."),
        ("7 năng lực đặc thù", "Năng lực ngôn ngữ", 3, "Có nhóm chữ cái – từ ngữ – từ vựng, bảng chữ cái đã việt hóa đủ dấu. Nhưng bộ từ theo âm tiết chưa soạn nên chưa đủ dùng."),
        ("7 năng lực đặc thù", "Năng lực tính toán", 5, "107 hoạt động toán: bảng cộng trừ nhân chia, số thập phân, phân số, đo lường, đặt tính dọc, phần bù của 10."),
        ("7 năng lực đặc thù", "Năng lực khoa học", 5, "Mạch điện tương tự và Mạch điện số là mô phỏng thật có tính toán dòng – áp; thêm Vòng tuần hoàn nước, Trọng lực, Hệ Mặt Trời."),
        ("7 năng lực đặc thù", "Năng lực công nghệ", 5, "Lập trình đường đi có thủ tục và vòng lặp; Mã hoá – giải mã đường đi; Bóng đèn nhị phân; Tàu ngầm."),
        ("7 năng lực đặc thù", "Năng lực tin học", 5, "Đúng vai trò nòng cốt theo CV 3456: làm quen chuột, bàn phím, soạn thảo, lập trình khối lệnh đầu tiên."),
        ("7 năng lực đặc thù", "Năng lực thẩm mĩ", 5, "Vẽ tự do, Vẽ bằng bánh răng, Sáng tác dương cầm, Chơi dương cầm, Ghép tranh danh họa."),
        ("7 năng lực đặc thù", "Năng lực thể chất", 2, "Không có. Bù bằng NeoArcade và NeoAiSport."),
    ]
    r = rows(ws, r + 1, d, level_col=3, bold_col=2)
    widths(ws, {"A": 20, "B": 30, "C": 11, "D": 86})

    # --------------------------------------------------------------- NLS
    ws = wb.create_sheet("NLS-TT02-2025")
    title(ws, "ĐỐI CHIẾU KHUNG NĂNG LỰC SỐ CHO NGƯỜI HỌC — TT 02/2025/TT-BGDĐT",
          "GCompris tiếng Việt (ThingEdu) — 6 miền • 24 năng lực thành phần • bậc mục tiêu Tiểu học 1–2", 6)
    r = legend(ws, 4, 6)
    header(ws, r, ["Miền năng lực", "Mã", "Năng lực thành phần", "Mức liên kết",
                   "Bậc mục tiêu (Tiểu học)", "Minh chứng trong bản GCompris tiếng Việt"])
    d = [
        ("I. Khai thác dữ liệu và thông tin", "1.1", "Duyệt, tìm kiếm và lọc dữ liệu, thông tin, nội dung số", 3, "1", "Menu 202 hoạt động có ô tìm kiếm và lọc theo nhóm, theo độ khó — trẻ tự tìm hoạt động mình cần. Chưa phải tìm tin trên Internet."),
        ("I. Khai thác dữ liệu và thông tin", "1.2", "Đánh giá dữ liệu, thông tin và nội dung số", 3, "1", "Đọc trục số có vạch chia, So sánh số, Câu hỏi trắc nghiệm — đọc và đối chiếu dữ liệu. Chưa có nội dung phân biệt tin thật – tin giả."),
        ("I. Khai thác dữ liệu và thông tin", "1.3", "Quản lý dữ liệu, thông tin và nội dung số", 3, "1", "Lưu và mở lại bài làm của mình ở Vẽ tự do, Sáng tác dương cầm, Hộp thăng bằng, Soạn thảo cho bé."),
        ("II. Giao tiếp và hợp tác số", "2.1", "Tương tác thông qua công nghệ số", 3, "1", "Chế độ máy chủ của thầy cô: học sinh đăng nhập, nhận bài và gửi kết quả về máy giáo viên."),
        ("II. Giao tiếp và hợp tác số", "2.2", "Chia sẻ thông qua công nghệ số", 2, "—", "Không có chức năng chia sẻ. Việc này thuộc Thing Notebook."),
        ("II. Giao tiếp và hợp tác số", "2.3", "Tham gia với tư cách công dân qua công nghệ số", 2, "—", "Ngoài phạm vi phần mềm."),
        ("II. Giao tiếp và hợp tác số", "2.4", "Hợp tác thông qua công nghệ số", 3, "1", "23 hoạt động hai người chơi trên cùng một máy. Hợp tác qua mạng thì không có."),
        ("II. Giao tiếp và hợp tác số", "2.5", "Nghi thức số", 2, "—", "Không có nội dung. Mentor phải đặt quy tắc ứng xử khi chơi chung máy."),
        ("II. Giao tiếp và hợp tác số", "2.6", "Quản lý danh tính số", 2, "—", "Chỉ có tên đăng nhập lớp học, không có nội dung về danh tính số."),
        ("III. Sáng tạo nội dung số", "3.1", "Phát triển nội dung số", 5, "1–2", "Vẽ tự do (đủ bộ cọ, hình học, chuyển sắc, đóng dấu, chữ), Vẽ bằng bánh răng, Sáng tác dương cầm, Soạn thảo cho bé — trẻ tạo ra tệp của riêng mình."),
        ("III. Sáng tạo nội dung số", "3.2", "Tích hợp và tinh chỉnh nội dung số", 3, "1", "Mở lại bản vẽ, bản nhạc đã lưu để sửa tiếp. Trình soạn màn chơi của Hộp thăng bằng cho trẻ tự làm màn rồi thử."),
        ("III. Sáng tạo nội dung số", "3.3", "Bản quyền và giấy phép", 3, "1", "Bản thân phần mềm là AGPL v3, tranh dùng có ghi nguồn Wikimedia Commons. Là dịp Mentor nói về nguồn mở, nhưng phần mềm không dạy điều đó."),
        ("III. Sáng tạo nội dung số", "3.4", "Lập trình", 5, "1–2", "Lập trình đường đi có lệnh, thủ tục và vòng lặp — bậc thang đầu tiên trước khi sang thẻ lệnh và Blockly. Thêm Mã hoá – giải mã đường đi, Mạch điện số."),
        ("IV. An toàn", "4.1", "Bảo vệ thiết bị", 2, "—", "Không có nội dung. Thuộc quy trình mượn – trả thiết bị của Làng."),
        ("IV. An toàn", "4.2", "Bảo vệ dữ liệu cá nhân và quyền riêng tư", 2, "—", "Không có. GCompris chạy hoàn toàn ngoại tuyến, không thu dữ liệu trẻ — đây là ưu điểm về quyền riêng tư nhưng không phải nội dung dạy học."),
        ("IV. An toàn", "4.3", "Bảo vệ sức khỏe và tinh thần", 3, "1", "Không có bài riêng, nhưng phần mềm không quảng cáo, không mua bán trong ứng dụng, không xếp hạng gây áp lực. Thời lượng màn hình do Mentor kiểm soát."),
        ("IV. An toàn", "4.4", "Bảo vệ môi trường", 3, "1", "Năng lượng tái tạo và Vòng tuần hoàn nước dạy hệ thống điện sạch và tài nguyên nước."),
        ("V. Giải quyết vấn đề", "5.1", "Giải quyết các vấn đề kỹ thuật", 5, "1–2", "Mạch điện tương tự và Mạch điện số: trẻ tự tìm ra vì sao đèn không sáng, sửa chập, thay linh kiện cháy. Có cả cảnh báo vòng nguồn điện áp."),
        ("V. Giải quyết vấn đề", "5.2", "Xác định nhu cầu và giải pháp công nghệ", 3, "1", "Trẻ tự chọn hoạt động và cấp độ hợp với mình. Xác định nhu cầu thật thì thuộc The Lab."),
        ("V. Giải quyết vấn đề", "5.3", "Sử dụng sáng tạo công nghệ số", 5, "1–2", "Trình soạn màn chơi Hộp thăng bằng, Vẽ tự do, Sáng tác dương cầm — dùng máy để làm ra cái mới chứ không chỉ trả lời câu hỏi."),
        ("V. Giải quyết vấn đề", "5.4", "Xác định thiếu hụt về năng lực số", 3, "1", "Thang sao vàng – sao đỏ và cấp độ trong từng hoạt động cho trẻ tự thấy mình đang ở đâu."),
        ("VI. Ứng dụng trí tuệ nhân tạo", "6.1", "Hiểu biết về AI", 2, "—", "Không có nội dung AI. Thuộc The Lab và NeoAiSport."),
        ("VI. Ứng dụng trí tuệ nhân tạo", "6.2", "Sử dụng AI có đạo đức, trách nhiệm", 2, "—", "Không có nội dung AI."),
        ("VI. Ứng dụng trí tuệ nhân tạo", "6.3", "Đánh giá và sử dụng công cụ AI phù hợp", 2, "—", "Không có nội dung AI. Lưu ý: kho giọng đọc tiếng Việt của bản này do AI sinh (VieNeu-TTS) — là sản phẩm của AI chứ không phải bài học về AI."),
    ]
    r = rows(ws, r + 1, d, level_col=4, bold_col=3)
    ws.cell(r + 1, 1, "Đọc thẳng: GCompris mạnh ở miền III (sáng tạo nội dung, lập trình) và miền V (giải quyết vấn đề). "
                      "Miền IV (an toàn) và miền VI (AI) gần như trống — đừng hứa hai miền này bằng GCompris.").font = Font(name="Arial", size=9, bold=True, color="FFC8402F")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    widths(ws, {"A": 28, "B": 7, "C": 34, "D": 11, "E": 14, "F": 74})

    # ------------------------------------------------------- CV 3456
    ws = wb.create_sheet("CV3456-TrienKhai")
    title(ws, "ĐỐI CHIẾU CÔNG VĂN 3456/BGDĐT-GDPT NGÀY 27/6/2025",
          "Hướng dẫn triển khai khung năng lực số cho học sinh phổ thông từ năm học 2025–2026 — "
          "GCompris tiếng Việt phục vụ được đến đâu", 4)
    r = legend(ws, 4, 4)
    header(ws, r, ["Yêu cầu / con đường triển khai theo CV 3456", "Mức liên kết",
                   "GCompris tiếng Việt đáp ứng thế nào", "Việc Mentor và nhà trường phải làm thêm"])
    d = [
        ("Không xây dựng môn học riêng, lồng ghép năng lực số vào các môn học hiện có", 5,
         "GCompris vốn được xếp theo môn: Toán, Tiếng Việt, Khoa học, Địa lí, Mĩ thuật, Âm nhạc, Tin học. "
         "Giáo viên mở đúng hoạt động của bài đang dạy, không phải dạy thêm tiết nào.",
         "Chọn sẵn hoạt động cho từng bài trong kế hoạch dạy học — dùng sheet HoatDong-202 để tra."),
        ("Môn Tin học giữ vai trò nòng cốt", 5,
         f"{n.get('Tin học', 0)} hoạt động thuộc nhóm Tin học: làm quen chuột, bàn phím, soạn thảo, và Lập trình đường đi "
         "có lệnh – thủ tục – vòng lặp. Đủ cho mạch 'Giải quyết vấn đề với sự trợ giúp của máy tính' ở tiểu học.",
         "Nối tiếp lên thẻ lệnh không màn hình rồi Blockly theo lộ trình The Lab."),
        ("Tích hợp qua hoạt động giáo dục, trải nghiệm, câu lạc bộ", 5,
         "Nhóm trò chơi tư duy (cờ vua, cờ đam, Tangram, Tháp Hà Nội, Sudoku, mê cung) hợp với giờ câu lạc bộ. "
         "Có sẵn chế độ hai người chơi, chơi cặp được ngay.",
         "Tổ chức thành câu lạc bộ có luân phiên máy, ghép cặp, và buổi Chia sẻ cuối."),
        ("Không gây quá tải cho giáo viên và học sinh", 5,
         "Phần mềm chạy ngoại tuyến, không cần tài khoản, không cần Internet, mở là chơi. "
         "Giáo viên không phải soạn học liệu số mới.",
         "Giới hạn thời lượng màn hình — khuyến nghị của ThingEdu là tối đa 15–20 phút mỗi buổi, "
         "phần còn lại làm tay."),
        ("Phù hợp Chương trình GDPT 2018", 5,
         "Xem sheet GDPT2018-MonHoc và GDPT2018-NangLuc: mức 5 ở Toán, Khoa học, Tin học, Mĩ thuật, "
         "Âm nhạc, năng lực tính toán, khoa học, công nghệ, tin học, thẩm mĩ.",
         "Riêng Tiếng Việt và Lịch sử – Địa lí còn thiếu dữ liệu Việt Nam, chưa dùng thay bài học được."),
        ("Khuyến khích ứng dụng AI, IoT, VR", 3,
         "GCompris không có AI, IoT hay VR trong nội dung. Nhưng kho giọng đọc tiếng Việt của bản này "
         "do AI sinh (VieNeu-TTS chạy ngoại tuyến trên máy) — bản thân việc đó là một câu chuyện AI kể được cho học sinh.",
         "Phần AI – IoT thật thuộc The Lab, ThingBot và NeoAiSport, không hứa bằng GCompris."),
        ("Bảo đảm học liệu số và hạ tầng", 5,
         "Toàn bộ 202 hoạt động, bản dịch và kho giọng đóng gói chạy offline trên NEO One (Linux ARM64), "
         "không phụ thuộc đường truyền. Giấy phép AGPL v3 — trường dùng và sao chép tự do, không phí bản quyền.",
         "Cấu hình máy theo deploy/install_vi.sh và tắt tự động tải dữ liệu ngoài."),
        ("Bảo đảm an toàn cho học sinh trên môi trường số", 5,
         "Không quảng cáo, không mua bán trong ứng dụng, không thu thập dữ liệu, không kết nối ra ngoài "
         "khi đã tắt tự động tải. Đây là điểm mạnh khi làm việc với phụ huynh.",
         "Nội dung DẠY về an toàn số thì phần mềm không có — xem miền IV sheet NLS."),
        ("Kiểm tra, đánh giá năng lực số của học sinh", 3,
         "Có chế độ máy chủ của thầy cô: giao bộ hoạt động, học sinh đăng nhập, kết quả gửi về máy giáo viên. "
         "Mỗi hoạt động có cấp độ và sao độ khó làm thang tự đánh giá.",
         "Chưa xuất được báo cáo theo mã chỉ báo NLS của CV 3456. Phải ghi tay vào Thing Notebook."),
        ("Dùng bảng mã chỉ báo năng lực số theo cấp học", 3,
         "Sheet NLS-TT02-2025 của file này đã gắn từng năng lực thành phần với hoạt động cụ thể, "
         "bậc mục tiêu Tiểu học 1–2.",
         "Khi phát hành tài liệu chính thức, đối chiếu nguyên văn tên 24 năng lực thành phần "
         "trong Phụ lục TT 02/2025 và bảng mã chỉ báo kèm CV 3456."),
    ]
    r = rows(ws, r + 1, d, level_col=2, bold_col=1)
    widths(ws, {"A": 44, "B": 11, "C": 66, "D": 56})

    # ------------------------------------------------------- Papert-Maker
    ws = wb.create_sheet("Papert-Maker")
    title(ws, "LĂNG KÍNH PAPERT VÀ MAKER — 202 HOẠT ĐỘNG ĐƯỢC DÙNG THẾ NÀO CHO ĐÚNG",
          "Constructionism: trẻ học tốt nhất khi tự tay làm ra một thứ có ý nghĩa với mình", 5)
    r = 4
    for t, b in [
        ("Vì sao phải phân loại",
         "Papert phản đối chính cái mà phần mềm giáo dục hay làm nhất: máy hỏi – trẻ đáp – máy chấm. "
         "Ông gọi đó là 'đặt máy tính vào chỗ của quyển vở'. Cái ông muốn là vi thế giới: một thế giới nhỏ có luật "
         "riêng, trẻ thò tay vào nghịch, đoán, thử, rồi tự thấy mình sai ở đâu. "
         "GCompris có cả hai loại. Bảng này tách chúng ra để Mentor biết hoạt động nào chơi là đủ, "
         "hoạt động nào bắt buộc phải nối sang việc làm tay."),
        ("Bốn vai trò dùng trong bảng", ""),
    ]:
        c = ws.cell(r, 1, t); c.font = Font(name="Arial", size=10, bold=True, color=BLUE)
        if b:
            cb = ws.cell(r, 2, b); cb.font = Font(name="Arial", size=9)
            cb.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            ws.row_dimensions[r].height = 60
        r += 1
    by_role = {}
    for a in acts:
        by_role.setdefault(role_of(a), []).append(a)
    header(ws, r, ["Vai trò trong tư duy Papert", "Số hoạt động", "Ý nghĩa",
                   "Hoạt động tiêu biểu (tên tiếng Việt)", "Cách Mentor dùng"])
    use = {
        "Vi thế giới": "Để trẻ nghịch tự do trước khi giảng. Không đặt câu hỏi đúng–sai. "
                       "Hỏi 'con thử đoán xem chuyện gì xảy ra nếu…' rồi để trẻ tự kiểm chứng.",
        "Công cụ sáng tạo": "Giao một sản phẩm để làm, không giao bài tập. In hoặc chiếu sản phẩm của trẻ "
                            "trong buổi Chia sẻ.",
        "Đồ vật để suy nghĩ": "Cho chơi cặp, rồi bắt trẻ nói ra chiến thuật của mình. "
                              "Việc nói ra mới là chỗ học, không phải việc thắng.",
        "Luyện tập có phản hồi": "Tối đa 10 phút. BẮT BUỘC nối sang việc làm tay cùng nội dung "
                                 "(xem cột 'Nối tay' ở sheet HoatDong-202). Nếu chỉ dùng nhóm này thì "
                                 "GCompris chỉ còn là quyển vở bài tập điện tử.",
    }
    d = []
    for role in ["Vi thế giới", "Công cụ sáng tạo", "Đồ vật để suy nghĩ", "Luyện tập có phản hồi"]:
        items = by_role.get(role, [])
        ex = " • ".join(x["vi_title"] for x in items[:6] if x["vi_title"])
        d.append((role, len(items), ROLE_NOTE[role], ex, use[role]))
    r = rows(ws, r + 1, d, bold_col=1)
    r += 1
    note = ws.cell(r + 1, 1,
                   f"Đọc thẳng: {len(by_role.get('Luyện tập có phản hồi', []))}/{len(acts)} hoạt động là luyện tập. "
                   "Đó là bản chất của GCompris và không có gì sai — nhưng nếu nhà trường chỉ mở nhóm đó thì "
                   "phần mềm không mang lại điều Papert nói tới. Giá trị Maker nằm ở "
                   f"{len(by_role.get('Vi thế giới', [])) + len(by_role.get('Công cụ sáng tạo', []))} hoạt động "
                   "vi thế giới và công cụ sáng tạo, cộng với việc nối tay sau mỗi buổi.")
    note.font = Font(name="Arial", size=9, bold=True, color="FFC8402F")
    note.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
    ws.row_dimensions[r + 1].height = 46
    r += 3
    c = ws.cell(r, 1, "Bốn nguyên tắc Papert áp vào buổi học GCompris")
    c.font = Font(name="Arial", size=10, bold=True, color=BLUE)
    r += 1
    header(ws, r, ["Nguyên tắc", "Nghĩa là gì", "Làm gì trong buổi có GCompris", "Dấu hiệu làm sai", ""])
    d = [
        ("Sàn thấp – trần cao – tường rộng", "Vào dễ, đi xa được, và có nhiều hướng đi khác nhau.",
         "Cho trẻ tự chọn hoạt động trên menu thay vì giao đồng loạt. Thang sao vàng – sao đỏ chính là cái trần.",
         "Cả lớp mở cùng một hoạt động, cùng một cấp độ, trong cùng số phút.", ""),
        ("Gỡ lỗi thay vì chấm sai", "Sai không phải là điểm trừ mà là dữ liệu để sửa.",
         "Khi trẻ sai ở Mạch điện hay Lập trình đường đi, hỏi 'con nghĩ nó hỏng ở đâu?' rồi im lặng chờ.",
         "Mentor chỉ ngay đáp án, hoặc bấm hộ trẻ.", ""),
        ("Đồ vật để suy nghĩ", "Ý tưởng trừu tượng phải có một vật cầm nắm được đi kèm.",
         "Mỗi hoạt động màn hình đi với một vật thật: xúc xắc, que tính, cân đĩa, mạch điện, bộ Tangram bìa.",
         "Buổi học chỉ có màn hình, bàn không có gì để sờ.", ""),
        ("Sản phẩm có người xem", "Trẻ làm ra thứ để khoe, không phải để nộp.",
         "Bản vẽ ở Vẽ tự do, bản nhạc ở Sáng tác dương cầm, màn chơi tự làm ở Hộp thăng bằng — "
         "in ra, chiếu lên, dán tường trong buổi Chia sẻ.",
         "Sản phẩm của trẻ nằm im trong máy, hết buổi là mất.", ""),
    ]
    r = rows(ws, r + 1, d, bold_col=1)
    widths(ws, {"A": 30, "B": 40, "C": 60, "D": 42, "E": 44})

    # ---------------------------------------------------- ma trận 202 hoạt động
    ws = wb.create_sheet("HoatDong-202")
    title(ws, "MA TRẬN 202 HOẠT ĐỘNG — GCOMPRIS TIẾNG VIỆT",
          "Tên tiếng Việt lấy đúng bản dịch đang chạy • Vai trò Papert và cột Nối tay là đề xuất của ThingEdu", 9)
    r = 4
    header(ws, r, ["#", "Tên hoạt động (tiếng Việt)", "Mã hoạt động", "Độ khó (sao)",
                   "Độ tuổi gợi ý", "Môn GDPT 2018", "Năng lực số (TT 02/2025)",
                   "Vai trò Papert", "Nối tay (việc làm thật đi kèm)"])
    order = {"Vi thế giới": 0, "Công cụ sáng tạo": 1, "Đồ vật để suy nghĩ": 2, "Luyện tập có phản hồi": 3}
    acts_sorted = sorted(acts, key=lambda a: (order[role_of(a)], a["diff"], a["vi_title"]))
    d = []
    for i, a in enumerate(acts_sorted, 1):
        d.append((i, a["vi_title"] or a["en_title"], a["name"],
                  ("★" * a["diff"]) or "—", AGE.get(a["diff"], "—"),
                  subject_of(a), nls_of(a), role_of(a), maker_of(a)))
    r = rows(ws, r + 1, d, bold_col=2)
    widths(ws, {"A": 5, "B": 34, "C": 24, "D": 10, "E": 20, "F": 30,
                "G": 30, "H": 20, "I": 52})
    ws.auto_filter.ref = f"A4:I{r - 1}"

    # ------------------------------------------------------ cần thiết kế lại
    ws = wb.create_sheet("CanThietKeLai")
    title(ws, "BỐN CHỖ KHÔNG DỊCH ĐƯỢC — PHẢI THIẾT KẾ LẠI CHO VIỆT NAM",
          "Đọc trước khi đưa GCompris tiếng Việt vào lớp học", 5)
    r = 4
    header(ws, r, ["Nội dung", "Hoạt động liên quan", "Vấn đề", "Hệ quả nếu dùng nguyên trạng",
                   "Việc phải làm"])
    d = [
        ("Xưng hô gia đình", "Gia đình · Chỉ đúng người thân",
         "Hoạt động tự mô tả là dạy quan hệ họ hàng 'theo hệ tuyến tính dùng ở phần lớn xã hội phương Tây'. "
         "Cây gia đình gốc có ĐÚNG MỘT ô Uncle và MỘT ô Aunt. Tiếng Việt tách thành chú, bác, cậu, cô, dì, "
         "thím, mợ theo bên nội – ngoại và thứ bậc tuổi; anh/chị/em còn phụ thuộc tuổi.",
         "DẠY SAI. Trẻ học rằng em trai của bố và anh của mẹ gọi giống nhau.",
         "Hoặc dựng lại bộ dữ liệu cây gia đình cho đúng cách xưng hô Việt (thêm nhánh nội – ngoại, "
         "thêm dấu hiệu tuổi), hoặc BỎ hai hoạt động này khỏi bản dùng trong lớp. "
         "Bản dịch hiện tại tạm theo bên nội."),
        ("Tiền", "Tiền · Tiền có xu · Thối tiền cho Tux",
         "Dùng ảnh đồng euro. Việt Nam không tiêu tiền xu, mệnh giá là tờ từ 1.000 tới 500.000 đồng.",
         "Trẻ luyện đếm tiền bằng đơn vị không tồn tại ở Việt Nam.",
         "Vẽ lại bộ ảnh mệnh giá tiền Việt và sửa dữ liệu. LƯU Ý PHÁP LÝ: mô phỏng hình ảnh tiền đồng "
         "có quy định riêng của Ngân hàng Nhà nước — kiểm tra trước khi phát hành công khai."),
        ("CHỦ QUYỀN — bản đồ Việt Nam thiếu Hoàng Sa và Trường Sa",
         "Tìm quốc gia trên bản đồ (cấp 12 Đông Nam Á)",
         "Tệp vietnam.svgz chỉ vẽ phần đất liền và vài đảo ven bờ. KHÔNG có quần đảo Hoàng Sa, "
         "KHÔNG có quần đảo Trường Sa. Nền bản đồ Đông Nam Á cũng để Biển Đông hoàn toàn trống. "
         "Điểm tích cực: bản đồ Trung Quốc không có đường lưỡi bò.",
         "VI PHẠM Nghị định 18/2020/NĐ-CP Điều 11 khoản 2: lưu hành sản phẩm bản đồ liên quan chủ "
         "quyền mà không thể hiện đúng chủ quyền bị phạt 30-40 triệu đồng, tịch thu tang vật và "
         "buộc cải chính. Tuyệt đối không đưa vào lớp học ở dạng hiện tại.",
         "1) Khoá ngay hai hoạt động bản đồ khỏi bản dùng trong trường. 2) Vẽ lại vietnam.svgz và "
         "nền southeast_asia.svgz có đủ hai quần đảo, kèm nhãn. 3) Gửi bản vá lên KDE. "
         "Xem DOCS/LUU_Y_BAN_DO_CHU_QUYEN.md."),
        ("Bản đồ hành chính Việt Nam", "Tìm vùng trên bản đồ",
         "18 bộ bản đồ hành chính: Ý, Ấn Độ, Trung Quốc, Úc, Mỹ, Pháp, Đức, Scotland, Romania, "
         "Litva… KHÔNG có Việt Nam.",
         "Không dùng được cho phần Địa lí Việt Nam của môn Lịch sử và Địa lí.",
         "Bổ sung bộ bản đồ 34 tỉnh thành. GCompris cho thêm bộ bản đồ mới mà không phải sửa mã nguồn."),
        ("Đài Loan liệt ngang hàng quốc gia", "Tìm quốc gia trên bản đồ (cấp 13 Đông Á)",
         "Cấp Đông Á xếp Đài Loan thành một mảnh riêng, ngang với Trung Quốc, Nhật Bản, Hàn Quốc.",
         "Việt Nam theo chính sách Một Trung Quốc; sách giáo khoa Việt Nam không liệt Đài Loan "
         "là quốc gia.",
         "Nhà trường cân nhắc: bỏ cấp Đông Á, hoặc giữ và giải thích. Cần chủ dự án chốt."),
        ("Bộ từ theo cấp độ âm tiết", "Chữ cái rơi · Từ rơi · Trò đoán chữ · Bấm vào chữ cái",
         "Các hoạt động này xây trên giả định 'chữ cái rời ghép thành từ'. Tiếng Việt có 29 chữ cái, "
         "thêm ă â đ ê ô ơ ư, 5 dấu thanh nằm trên nguyên âm, và đơn vị đọc là âm tiết.",
         "Trẻ gõ chữ cái rời theo lối tiếng Anh, không khớp cách học vần lớp 1.",
         "Soạn bộ từ default-vi.json theo cấp độ âm tiết cho Chữ cái rơi và Từ rơi. "
         "Bảng chữ cái tiếng Việt đã việt hóa xong cho Thứ tự bảng chữ cái và Sắp chữ cái theo thứ tự."),
        ("(bổ sung) Kho giọng đọc chưa đủ", "Mở rộng vốn từ · các hoạt động địa lí",
         "Mới sinh 202/888 tệp giọng. Còn thiếu 564 từ vựng và 129 tên nước.",
         "Hoạt động Mở rộng vốn từ chưa dùng được; trẻ chưa biết đọc sẽ mất phần nghe.",
         "Sinh nốt bằng VieNeu-TTS sau khi dịch content-vi.json (1.090 từ)."),
    ]
    r = rows(ws, r + 1, d, bold_col=1)
    widths(ws, {"A": 24, "B": 30, "C": 60, "D": 40, "E": 62})

    # --------------------------------------------------------------- TongHop
    ws = wb.create_sheet("TongHop", 1)
    title(ws, "TỔNG HỢP MỨC LIÊN KẾT THEO TỪNG HỆ CHUẨN", "GCompris tiếng Việt (ThingEdu)", 5)
    r = 4
    header(ws, r, ["Hệ chuẩn (sheet)", "Số chuẩn mức 5", "Số chuẩn mức 3", "Số chuẩn mức 2", "Đọc nhanh"])
    def dem(sheet, col):
        c = {2: 0, 3: 0, 5: 0}
        for row in wb[sheet].iter_rows(min_row=9):
            v = row[col - 1].value
            if v in c:
                c[v] += 1
        return c[5], c[3], c[2]

    g5, g3, g2 = dem("GDPT2018-MonHoc", 2)
    n5, n3, n2 = dem("GDPT2018-NangLuc", 3)
    s5, s3, s2 = dem("NLS-TT02-2025", 4)
    c5, c3, c2 = dem("CV3456-TrienKhai", 2)
    d = [
        ("GDPT 2018 — Môn học & HĐGD", g5, g3, g2,
         "Mạnh ở Toán, Khoa học, Tin học, Mĩ thuật, Âm nhạc, HĐTN. Yếu ở Tiếng Việt và Lịch sử – Địa lí "
         "vì thiếu dữ liệu Việt Nam. Trống ở Đạo đức và Thể chất."),
        ("GDPT 2018 — Phẩm chất & Năng lực", n5, n3, n2,
         "Mạnh ở năng lực tính toán, khoa học, công nghệ, tin học, thẩm mĩ và tự chủ – tự học."),
        ("Khung năng lực số TT 02/2025", s5, s3, s2,
         "Mạnh ở miền III (sáng tạo nội dung, lập trình) và miền V (giải quyết vấn đề kỹ thuật). "
         "Miền IV (an toàn) và VI (AI) gần như trống — không hứa hai miền này bằng GCompris."),
        ("Công văn 3456/BGDĐT-GDPT", c5, c3, c2,
         "Đáp ứng tốt cả bốn con đường triển khai mà công văn nêu, và đặc biệt hợp với yêu cầu "
         "'không gây quá tải' vì chạy offline, không cần tài khoản, giáo viên không phải soạn học liệu mới."),
    ]
    r = rows(ws, r + 1, d, bold_col=1)
    r += 1
    for txt, color in [
        ("Đọc nhanh: mức 5 = lập luận chủ lực khi làm việc với trường và phòng giáo dục; "
         "mức 3 = cơ hội nâng cấp giáo án, giao Mentor; mức 2 = minh bạch phạm vi, không hứa quá.", "FF666666"),
        ("Một câu tóm cả file: GCompris tiếng Việt là công cụ luyện tập và khám phá trên màn hình, "
         "phục vụ tốt môn Toán – Khoa học – Tin học và nhiệm vụ năng lực số theo CV 3456. "
         "Nó KHÔNG phải chương trình Maker. Giá trị Maker chỉ xuất hiện khi mỗi buổi có nối tay.", "FFC8402F"),
    ]:
        c = ws.cell(r + 1, 1, txt)
        c.font = Font(name="Arial", size=9, bold=color != "FF666666", color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)
        ws.row_dimensions[r + 1].height = 32
        r += 2
    widths(ws, {"A": 36, "B": 15, "C": 15, "D": 15, "E": 76})
    return wb


if __name__ == "__main__":
    acts = json.load(open(sys.argv[1], encoding="utf-8"))
    wb, acts, cnt = build(acts, sys.argv[2])
    wb = build_rest(wb, acts, cnt)
    wb.save(sys.argv[2])
    print(f"đã ghi {sys.argv[2]} — {len(wb.sheetnames)} sheet: {', '.join(wb.sheetnames)}")
