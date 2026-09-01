#!/usr/bin/env python3
"""Danh sách 202 hoạt động GCompris tiếng Việt kèm ý nghĩa giáo dục.

Mỗi hoạt động đọc theo ba lăng kính: tư duy Papert, khung GDPT 2018, và khung
năng lực số TT 02/2025 (CV 3456 hướng dẫn thực hiện).

Usage: make_xlsx_hoatdong.py <acts.json> <out.xlsx>
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

sys.path.insert(0, __file__.rsplit("/", 1)[0])
from make_xlsx import (AGE, BORDER, BLUE, GREY, ROW_BG, header, maker_of,  # noqa: E402
                       role_of, rows, subject_of, title, widths)

# Năng lực GDPT 2018 gắn với từng môn
NL = {
    "Toán": "NL tính toán · NL giải quyết vấn đề",
    "Tiếng Việt": "NL ngôn ngữ · NL giao tiếp",
    "Tự nhiên và Xã hội (TH) · Khoa học · KHTN": "NL khoa học · NL tìm hiểu tự nhiên",
    "Tin học": "NL tin học · NL công nghệ",
    "Lịch sử và Địa lí": "NL tìm hiểu xã hội · NL khoa học",
    "Mĩ thuật": "NL thẩm mĩ · NL sáng tạo",
    "Âm nhạc": "NL thẩm mĩ · NL cảm thụ âm nhạc",
    "Hoạt động trải nghiệm": "NL tự chủ và tự học · NL giải quyết vấn đề",
    "Hoạt động trải nghiệm (giáo dục hòa nhập)": "NL tự chủ · NL giao tiếp và hợp tác",
    "— (màn hình chính, không phải hoạt động học)": "NL tự chủ và tự học",
    "Tin học ": "NL tin học",
}

# Ý nghĩa Papert viết riêng cho những hoạt động đáng nói nhất
PAPERT = {
    "analog_electricity": "Vi thế giới điện đúng nghĩa Papert: mạch có luật vật lý riêng, trẻ nối sai thì đèn không sáng hoặc cháy, và phần mềm nói rõ vì sao. Sai ở đây là dữ liệu để sửa, không phải điểm trừ.",
    "digital_electricity": "Trẻ dựng mạch logic rồi thấy ngay đầu ra đổi theo. Khái niệm trừu tượng AND–OR–NOT có một vật cầm nắm được.",
    "programmingMaze": "Bản LOGO thu nhỏ: trẻ viết một dãy lệnh, chạy, thấy Tux đi sai rồi quay lại sửa lệnh. Có cả thủ tục và vòng lặp — hai ý tưởng lớn nhất mà Papert muốn trẻ gặp sớm.",
    "path_encoding": "Trẻ phải nghĩ thay cho Tux: mô tả một con đường bằng ký hiệu. Đây là bước đầu của việc lập trình — tách ý định ra khỏi hành động.",
    "path_decoding": "Đọc một dãy lệnh của người khác rồi làm theo — kỹ năng ngược của lập trình, cũng quan trọng ngang.",
    "path_encoding_relative": "Chuyển từ hướng cố định sang hướng theo thân mình. Papert gọi đây là 'body syntonic': trẻ đặt mình vào chỗ con rùa để nghĩ.",
    "path_decoding_relative": "Chuyển từ hướng cố định sang hướng theo thân mình — trẻ tưởng tượng mình là Tux mà quay trái, quay phải.",
    "gravity": "Trẻ cảm nhận lực hút bằng tay lái chứ không bằng công thức. Càng gần hành tinh, tay càng nặng — hiểu trước, đặt tên sau.",
    "land_safe": "Trò chơi này là một bài vật lý giấu kín: lực đẩy, gia tốc và trọng lực đều thật, trẻ học bằng cách hạ cánh hỏng vài chục lần.",
    "submarine": "Một hệ thống có nhiều van và nhiều biến cùng lúc. Trẻ phải xây mô hình trong đầu về cách các bộ phận ảnh hưởng nhau.",
    "watercycle": "Vi thế giới về một hệ thống mà trẻ thấy hằng ngày. Bấm đúng thứ tự thì Tux được tắm — mục tiêu cụ thể, có ý nghĩa với trẻ.",
    "renewable_energy": "Trẻ tự nối lại cả lưới điện. Ý tưởng lớn ở đây là ràng buộc hệ thống: không tiêu nhiều hơn lượng làm ra.",
    "solar_system": "Ở chế độ học, trẻ tự đi hỏi từng hành tinh. Đồng hồ Độ gần biến câu trả lời sai thành chỉ dẫn 'nóng – lạnh' thay vì dấu X.",
    "canal_lock": "Một cơ cấu thật của thế giới người lớn, thu nhỏ lại vừa tay trẻ. Trẻ phải hiểu nguyên lý mới mở đúng thứ tự van.",
    "balancebox": "Có trình soạn màn chơi: trẻ chuyển từ người chơi thành người làm ra trò chơi. Đây đúng là điều Papert mong nhất ở phần mềm giáo dục.",
    "binary_bulb": "Hệ nhị phân được làm thành tám bóng đèn bật tắt. Trẻ sờ được vào cách máy tính đếm.",
    "color_mix": "Trẻ kéo thanh trượt và thấy màu đổi ngay. Quy luật pha màu trừ tự hiện ra, không cần ai giảng.",
    "color_mix_light": "Cùng một cách học nhưng cho màu ánh sáng — và trẻ tự phát hiện nó ngược với pha màu sơn.",
    "scalesboard": "Cân thăng bằng là mô hình vật lý của dấu bằng. Trẻ hiểu phương trình bằng tay trước khi gặp chữ x.",
    "share": "Phép chia có dư được kể thành chuyện chia kẹo cho bạn. Phần dư không phải con số lạ mà là mấy cái kẹo còn lại trong lọ.",
    "learn_decimals": "Số thập phân được làm thành thanh và ô vuông. Trẻ nhìn thấy một phần mười là gì trước khi viết dấu phẩy.",
    "fractions_create": "Phân số hiện ra thành cái bánh chia phần. Tử số và mẫu số có nghĩa cụ thể chứ không phải hai con số chồng nhau.",
    "graduated_line_read": "Trục số thành một cái thước dài. Trẻ đọc vị trí trên đó như đọc thước kẻ thật.",
    "sketch": "Công cụ vẽ đầy đủ, không có bài tập nào. Trẻ làm ra bức tranh của mình, và bức tranh đó có thể in ra treo lên.",
    "drawing_wheels": "Bộ vẽ spirograph số. Trẻ đổi số răng rồi thấy hình đổi theo — quan hệ toán học hiện ra thành cái đẹp.",
    "compass": "Compa số có cả chế độ tự do lẫn mẫu để bắt chước. Hình học thành việc dựng hình chứ không phải học thuộc.",
    "piano_composition": "Trẻ viết ra bản nhạc của mình rồi nghe lại. Ký hiệu nhạc trở thành công cụ để nói điều mình muốn.",
    "baby_wordprocessor": "Trang giấy trắng đầu tiên trên máy. Không có câu hỏi nào — trẻ viết cái trẻ muốn.",
    "simplepaint": "Ô màu và bảng màu, không có đúng sai. Bước đầu tiên vào việc dùng máy để làm ra thứ của mình.",
    "chess": "Cờ vua là đồ vật để suy nghĩ kinh điển: mỗi nước đi là một giả thuyết, đối thủ là người phản biện.",
    "hanoi_real": "Bài toán đệ quy nổi tiếng, cầm nắm được. Trẻ tìm ra quy luật bằng tay trước khi ai nói tới từ 'đệ quy'.",
    "tangram": "Bảy mảnh ghép, vô số hình. Sàn thấp trần cao đúng nghĩa: ai cũng ghép được hình đầu, hình cuối thì khó thật.",
    "sudoku": "Suy luận loại trừ thuần túy. Không cần biết chữ, không cần biết tính — chỉ cần nghĩ.",
    "traffic": "Bài toán dọn đường: muốn xe đỏ ra thì phải nghĩ ngược từ đích. Trẻ học lập kế hoạch nhiều bước.",
    "superbrain": "Trẻ đặt giả thuyết, nhận phản hồi từng phần, rồi thu hẹp dần. Đây chính là phương pháp khoa học ở dạng trò chơi.",
    "lightsoff": "Mỗi cái bấm ảnh hưởng cả vùng xung quanh. Trẻ phải nghĩ về hệ thống chứ không phải từng bóng đèn.",
    "peg_solitaire": "Một mình với bàn cờ. Mỗi nước đi làm hẹp lựa chọn sau — bài học về hậu quả của quyết định.",
    "guesscount": "Trẻ tự dựng phép tính để ra kết quả cho trước. Ngược với bài tập thường ngày, và vì thế khó hơn nhiều.",
    "menu": "Không phải hoạt động học, nhưng quan trọng với Papert: trẻ tự chọn mình muốn làm gì, thay vì được giao bài.",
    "family": "CẨN THẬN: cây gia đình gốc theo lối phương Tây, chỉ có một ô 'chú/bác/cậu'. Đọc sheet CanThietKeLai trước khi dùng trong lớp.",
    "family_find_relative": "CẨN THẬN: cùng vấn đề xưng hô như hoạt động Gia đình. Chưa nên dùng để dạy.",
    "money": "CẨN THẬN: dùng đồng euro, không phải tiền Việt. Xem sheet CanThietKeLai.",
}

ROLE_FALLBACK = {
    "Vi thế giới": "Vi thế giới: trẻ thao tác trên một hệ có luật riêng và tự rút ra quy luật. {goal}",
    "Công cụ sáng tạo": "Công cụ để trẻ làm ra sản phẩm của mình, không có đáp án đúng sai. {goal}",
    "Đồ vật để suy nghĩ": "Bài toán cụ thể dùng làm chỗ dựa để nghĩ về chiến thuật và cấu trúc. {goal}",
    "Luyện tập có phản hồi": "Luyện tập: máy hỏi – trẻ đáp. {goal} Giữ tối đa 10 phút rồi nối sang việc làm tay.",
}


def papert_meaning(a):
    if a["name"] in PAPERT:
        return PAPERT[a["name"]]
    goal = (a["vi_goal"] or a["vi_desc"] or "").strip()
    return ROLE_FALLBACK[role_of(a)].format(goal=goal).replace("  ", " ").strip()


def nls_detail(a):
    from make_xlsx import nls_of
    base = nls_of(a)
    role = role_of(a)
    why = {
        "Vi thế giới": "Trẻ thử – sai – sửa trên máy, đúng mạch 'giải quyết vấn đề kỹ thuật' của miền V.",
        "Công cụ sáng tạo": "Trẻ tạo ra tệp của riêng mình, đúng mạch 'phát triển nội dung số' của miền III.",
        "Đồ vật để suy nghĩ": "Dùng máy như chỗ để nghĩ, thuộc mạch 'sử dụng sáng tạo công nghệ số'.",
        "Luyện tập có phản hồi": "Ở mức làm quen thiết bị và thao tác số cơ bản, bậc 1 của khung.",
    }[role]
    return f"{base} — {why}"


def build(acts, out_path):
    acts = [a for a in acts if a["name"] != "template"]
    wb = Workbook()

    ws = wb.active
    ws.title = "HuongDan"
    title(ws, "GCOMPRIS TIẾNG VIỆT — 202 HOẠT ĐỘNG VÀ Ý NGHĨA GIÁO DỤC",
          "ThingEdu · Làng Maker — đọc theo tư duy Papert, khung GDPT 2018 và khung năng lực số TT 02/2025", 2)
    txt = [
        ("File này để làm gì",
         "Danh sách đầy đủ 202 hoạt động của bản GCompris tiếng Việt, mỗi hoạt động kèm ý nghĩa giáo dục "
         "đọc theo ba lăng kính. Dùng khi soạn buổi học, khi chọn hoạt động cho một bài cụ thể, và khi "
         "giải thích với phụ huynh hay nhà trường vì sao cho trẻ chơi hoạt động đó."),
        ("Ba lăng kính", ""),
        ("Tư duy Papert",
         "Cột 'Vai trò Papert' và 'Ý nghĩa theo Papert'. Bốn vai trò: vi thế giới (trẻ nghịch một hệ có luật "
         "riêng), công cụ sáng tạo (trẻ làm ra sản phẩm), đồ vật để suy nghĩ (bài toán cụ thể để nghĩ), và "
         "luyện tập có phản hồi (máy hỏi – trẻ đáp). Papert xếp loại thứ tư là mức dùng máy tính yếu nhất, "
         "nên với những hoạt động đó cột 'Nối tay' là bắt buộc chứ không phải gợi ý."),
        ("Khung GDPT 2018",
         "Cột 'Môn GDPT 2018' và 'Năng lực'. Mục tiêu học tập lấy nguyên văn mục tiêu do nhóm tác giả "
         "GCompris viết, đã dịch sang tiếng Việt — không phải diễn giải của ThingEdu."),
        ("Khung năng lực số TT 02/2025",
         "Cột 'Năng lực số'. Mã năng lực thành phần theo Thông tư 02/2025/TT-BGDĐT, hướng dẫn thực hiện "
         "theo Công văn 3456/BGDĐT-GDPT ngày 27/6/2025. Bậc mục tiêu ở tiểu học là bậc 1–2."),
        ("Điều KHÔNG hứa",
         "Ba hoạt động có ghi CẨN THẬN ở cột ý nghĩa: Gia đình, Chỉ đúng người thân và Tiền. Nội dung gốc "
         "không hợp với Việt Nam — xem file đối chiếu chuẩn, sheet CanThietKeLai, trước khi đưa vào lớp."),
        ("Cột 'Nối tay'",
         "Việc làm bằng vật thật đi kèm hoạt động. Đây là chỗ GCompris chuyển từ phần mềm luyện tập "
         "thành một mảnh của chương trình Maker. Không có cột này thì buổi học chỉ còn màn hình."),
        ("Nguồn", "github.com/ThingEdu/gcompris-vi — GCompris 26.1, giấy phép AGPL v3, cộng đồng KDE. "
                  "Bản dịch 4.277 chuỗi, kho giọng tiếng Việt sinh bằng VieNeu-TTS. "
                  "Đã nghiệm thu chạy thật trên NEO One."),
    ]
    r = 4
    for a, b in txt:
        c = ws.cell(r, 1, a)
        c.font = Font(name="Arial", size=10, bold=True, color=BLUE if not b else "FF000000")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        if b:
            cb = ws.cell(r, 2, b)
            cb.font = Font(name="Arial", size=9)
            cb.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    widths(ws, {"A": 30, "B": 96})
    return wb, acts


def build_list(wb, acts):
    import collections

    ws = wb.create_sheet("202-HoatDong")
    title(ws, "202 HOẠT ĐỘNG — Ý NGHĨA THEO PAPERT, GDPT 2018 VÀ NĂNG LỰC SỐ",
          "Tên và mục tiêu lấy đúng bản dịch tiếng Việt đang chạy • "
          "Vai trò Papert, ý nghĩa và cột Nối tay là phân tích của ThingEdu", 13)
    r = 4
    header(ws, r, ["#", "Tên hoạt động", "Mã hoạt động", "Nhóm", "Sao", "Độ tuổi gợi ý",
                   "Việc trẻ làm (mô tả)", "Mục tiêu học tập (GDPT 2018)",
                   "Môn GDPT 2018", "Năng lực GDPT 2018",
                   "Năng lực số TT 02/2025 (CV 3456)", "Vai trò Papert",
                   "Ý nghĩa theo Papert", "Nối tay (việc làm thật)"])
    order = {"Vi thế giới": 0, "Công cụ sáng tạo": 1, "Đồ vật để suy nghĩ": 2,
             "Luyện tập có phản hồi": 3}
    acts_sorted = sorted(acts, key=lambda a: (order[role_of(a)], -a["diff"],
                                              a["vi_title"] or a["en_title"]))
    NHOM = {"math": "Toán", "arithmetic": "Số học", "sciences": "Khoa học",
            "experiment": "Thí nghiệm", "discovery": "Khám phá", "reading": "Đọc",
            "letters": "Chữ cái", "words": "Từ ngữ", "vocabulary": "Từ vựng",
            "computer": "Máy tính", "keyboard": "Bàn phím", "mouse": "Chuột",
            "memory": "Trí nhớ", "logic": "Logic", "strategy": "Chiến thuật",
            "puzzle": "Xếp hình", "fun": "Giải trí", "music": "Âm nhạc",
            "arts": "Mĩ thuật", "geography": "Địa lí", "history": "Lịch sử",
            "numeration": "Số", "measures": "Đo lường", "money": "Tiền",
            "maze": "Mê cung", "braille": "Chữ nổi", "color": "Màu sắc",
            "colors": "Màu sắc", "mobile": "Cảm ứng"}

    def nhom(a):
        for s in a["section"].split():
            if s in NHOM:
                return NHOM[s]
        return "Khác"

    data = []
    for i, a in enumerate(acts_sorted, 1):
        subj = subject_of(a)
        data.append((i, a["vi_title"] or a["en_title"], a["name"], nhom(a),
                     ("★" * a["diff"]) or "—", AGE.get(a["diff"], "—"),
                     a["vi_desc"] or "—", a["vi_goal"] or "—",
                     subj, NL.get(subj, "NL tự chủ và tự học"),
                     nls_detail(a), role_of(a), papert_meaning(a), maker_of(a)))
    r = rows(ws, r + 1, data, bold_col=2)
    widths(ws, {"A": 5, "B": 30, "C": 24, "D": 13, "E": 9, "F": 19, "G": 46,
                "H": 40, "I": 28, "J": 30, "K": 46, "L": 20, "M": 66, "N": 46})
    ws.auto_filter.ref = f"A4:N{r - 1}"

    # ------------------------------------------------------------- theo nhóm
    ws = wb.create_sheet("TheoNhom")
    title(ws, "TỔNG HỢP THEO NHÓM CHỦ ĐỀ VÀ VAI TRÒ PAPERT",
          "Dùng để chọn nhanh: buổi này cần nhóm nào, vai trò nào", 6)
    r = 4
    header(ws, r, ["Nhóm chủ đề", "Số hoạt động", "Vi thế giới", "Công cụ sáng tạo",
                   "Đồ vật để suy nghĩ", "Luyện tập có phản hồi"])
    by = collections.defaultdict(lambda: collections.Counter())
    tot = collections.Counter()
    for a in acts:
        by[nhom(a)][role_of(a)] += 1
        tot[nhom(a)] += 1
    d = [(k, tot[k], by[k]["Vi thế giới"], by[k]["Công cụ sáng tạo"],
          by[k]["Đồ vật để suy nghĩ"], by[k]["Luyện tập có phản hồi"])
         for k, _ in tot.most_common()]
    d.append(("TỔNG", sum(tot.values()),
              sum(x[2] for x in d), sum(x[3] for x in d),
              sum(x[4] for x in d), sum(x[5] for x in d)))
    r = rows(ws, r + 1, d, bold_col=1)
    c = ws.cell(r + 1, 1,
                "Cách đọc: cột 'Luyện tập có phản hồi' càng lớn thì nhóm đó càng cần nối tay. "
                "Nhóm Khoa học và Thí nghiệm có tỉ lệ vi thế giới cao nhất — đó là nhóm nên ưu tiên "
                "khi muốn buổi học mang tinh thần Papert.")
    c.font = Font(name="Arial", size=9, bold=True, color="FFC8402F")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    ws.row_dimensions[r + 1].height = 32
    widths(ws, {"A": 20, "B": 14, "C": 14, "D": 17, "E": 20, "F": 22})
    return wb


if __name__ == "__main__":
    acts = json.load(open(sys.argv[1], encoding="utf-8"))
    wb, acts = build(acts, sys.argv[2])
    wb = build_list(wb, acts)
    wb.save(sys.argv[2])
    print(f"đã ghi {sys.argv[2]} — {len(acts)} hoạt động, sheet: {', '.join(wb.sheetnames)}")
